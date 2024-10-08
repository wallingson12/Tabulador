import re
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import numpy as np
import cv2
from PyPDF2 import PdfReader
from pdf2image import convert_from_path
import pytesseract

# Configurar o caminho do Tesseract
caminho_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = caminho_tesseract

# Função para limpar texto removendo caracteres especiais
def clean_text(text):
    return re.sub(r'[^a-zA-Z0-9\s]', '', text)

# Função para limpar o CNPJ removendo caracteres extras (como vírgulas)
def clean_cnpj(cnpj):
    return re.sub(r'[^\d]', '', cnpj)

# Função para encontrar padrões de CNPJ, Nome, Data e Valores Monetários
def find_patterns(text):
    cnpj_pattern = re.compile(r'\b(\d{2}[.,\s]?\d{3}[.,\s]?\d{3}[./\s]{1,2}?\d{4}[-\s]?\d{2})\b')
    date_pattern = re.compile(r'(\d{2}/\d{2}/\d{4})')
    value_pattern = re.compile(r'(\d{1,3}(?:\.\d{3})*,\d{2})')

    matches = []
    lines = text.split('\n')
    cnpj, name, date = None, '', None
    values = []

    for line in lines:
        line = line.strip()
        cnpj_match = cnpj_pattern.search(line)
        if cnpj_match:
            if cnpj and name and date and values:
                matches.append((cnpj, name, date, values[0], values[1] if len(values) > 1 else ''))
            cnpj = clean_cnpj(cnpj_match.group(1))
            name = line[len(cnpj):].strip()
            date = None
            values = []

        if cnpj and not date:
            date_match = date_pattern.search(line)
            if date_match:
                date = date_match.group(1)

        if cnpj and date:
            value_matches = value_pattern.findall(line)
            if value_matches:
                values.extend(value_matches)

        if cnpj and date and any(value_pattern.search(part) for part in line.split()):
            if values:
                matches.append((cnpj, name, date, values[0], values[1] if len(values) > 1 else ''))
            cnpj, name, date, values = None, '', None, []

    if cnpj and name and date and values:
        matches.append((cnpj, name, date, values[0], values[1] if len(values) > 1 else ''))

    return matches

# Função para garantir que o nome contém apenas letras
def filter_name(name):
        return re.sub(r'[^A-Z\s]', '', name).strip()

# Função de pré-processamento de imagem usando binarização adaptativa, dilatação e erosão
def preprocess_image(image):
    img_np = np.array(image)
    gray = cv2.cvtColor(img_np, cv2.COLOR_BGR2GRAY)
    adaptive_thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    kernel = np.ones((2, 2), np.uint8)
    dilated = cv2.dilate(adaptive_thresh, kernel, iterations=1)
    eroded = cv2.erode(dilated, kernel, iterations=2)
    processed_img = Image.fromarray(eroded)
    return processed_img


# Função para extrair texto usando OCR
def extract_text_from_pdf_with_ocr(pdf_path, oem_mode=1):
    pages = convert_from_path(pdf_path, dpi=600)
    text = ''
    for page in pages:
        processed_image = preprocess_image(page)
        custom_config = f'--oem {oem_mode} --psm 4'
        text += pytesseract.image_to_string(processed_image, lang='por+eng', config=custom_config)
    return text

# Função para extrair texto de um PDF sem OCR
def extract_text_from_pdf(pdf_path):
    pdf = PdfReader(open(pdf_path, 'rb'))
    text = ''
    for page in pdf.pages:
        text += page.extract_text()
    return text

# Função para encontrar padrões de valores com OCR
def find_value_patterns(text):
    pattern = re.compile(r'(\d{4})\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})')
    date_pattern = re.compile(r'\d{2}/\d{2}/\d{4}')

    matches = []
    lines = text.split('\n')

    for line in lines:
        line = line.strip()
        value_match = pattern.search(line)
        if value_match:
            code, value1, value2 = value_match.groups()
            if not date_pattern.search(line):
                matches.append((code, value1, value2))

    return matches

# Função para extrair dados do PDF e salvar no Excel (para PDFs sem OCR)
def extract_data_to_excel(pdf_path, excel_path):
    text = extract_text_from_pdf(pdf_path)
    value_data = find_value_patterns(text)
    pattern_data = find_patterns(text)

    df_values = pd.DataFrame(value_data, columns=['Código', 'Rendimento', 'Imposto'])
    df_patterns = pd.DataFrame(
        [(cnpj, filter_name(name), date, val1, val2) for cnpj, name, date, val1, val2 in pattern_data],
        columns=['CNPJ', 'Nome', 'Data', 'Rendimento Tributável', 'Imposto Retido'])

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
            book = load_workbook(excel_path)

            if 'Valores' in book.sheetnames:
                std = book['Valores']
                book.remove(std)
            df_values.to_excel(writer, sheet_name='Valores', index=False)

            if 'Dados' in book.sheetnames:
                std = book['Dados']
                book.remove(std)
            df_patterns.to_excel(writer, sheet_name='Dados', index=False)

    except FileNotFoundError:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_values.to_excel(writer, sheet_name='Valores', index=False)
            df_patterns.to_excel(writer, sheet_name='Dados', index=False)

# Função para extrair dados do PDF e salvar no Excel com OCR
def extract_data_to_excel_with_ocr(pdf_path, excel_path, sheet_name):
    text = extract_text_from_pdf_with_ocr(pdf_path)
    data = find_patterns(text)
    print(text)
    data = [(cnpj, filter_name(name), date, val1, val2) for cnpj, name, date, val1, val2 in data]
    df = pd.DataFrame(data, columns=['CNPJ', 'Nome', 'Data', 'Rendimento Tributável', 'Imposto Retido'])

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                std = book[sheet_name]
                book.remove(std)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# Função para extrair valores e salvar no Excel com OCR
def extract_values_to_excel_with_ocr(pdf_path, excel_path, sheet_name):
    text = extract_text_from_pdf_with_ocr(pdf_path)
    data = find_value_patterns(text)
    df = pd.DataFrame(data, columns=['Código', 'Rendimento', 'Imposto'])

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
            book = load_workbook(excel_path)
            if sheet_name in book.sheetnames:
                std = book[sheet_name]
                book.remove(std)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# # Caminho do arquivo PDF de entrada e do arquivo Excel de saída
# pdf_path = r'1.20/Fontes Pagadoras - MTE Filial 0003-15.pdf'
# excel_path = r'1.20/Fontes Pagadoras - MTE Filial 0003-15.xlsx'

# EXCEL
# extract_data_to_excel(pdf_path, excel_path)

# OCR
# extract_data_to_excel_with_ocr(pdf_path, excel_path, 'Data')
# extract_values_to_excel_with_ocr(pdf_path, excel_path, 'Valores')
